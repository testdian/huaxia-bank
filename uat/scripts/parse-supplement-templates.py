#!/usr/bin/env python3
"""从人行48号文附2 Excel 生成 supplement-templates-data.js"""
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

ROOT = Path(__file__).resolve().parent.parent
PBO_PATH = Path(
    '/Users/fangdanyang/Desktop/华夏银行/人行投融资碳核算工作要求/'
    '附件5．中国人民银行信贷市场司关于开展年度金融机构投融资碳核算工作的通知（银信贷〔2026〕3号）/'
    '附件2 《操作指引》附2八大行业碳核算信息采集表、碳排放因子表.xlsx'
)

INDUSTRY_SHEETS = ['电力', '水泥', '平板玻璃', '钢铁', '铝冶炼', '铜冶炼', '石化', '化工', '造纸', '民航']

SHEET_META = {
    '电力': {'industryMajor': '电力', 'gbCodes': ['D4411', 'D4412', 'D4417', 'D4420']},
    '水泥': {'industryMajor': '建材', 'gbCodes': ['C3011']},
    '平板玻璃': {'industryMajor': '建材', 'gbCodes': ['C3041']},
    '钢铁': {'industryMajor': '钢铁', 'gbCodes': ['C3110', 'C3120', 'C3130']},
    '铝冶炼': {'industryMajor': '有色', 'gbCodes': ['C3216']},
    '铜冶炼': {'industryMajor': '有色', 'gbCodes': ['C3211']},
    '石化': {'industryMajor': '石化', 'gbCodes': ['C2511']},
    '化工': {'industryMajor': '化工', 'gbCodes': [
        'C2611', 'C2612', 'C2613', 'C2614', 'C2619',
        'C2621', 'C2622', 'C2623', 'C2624', 'C2625', 'C2629'
    ]},
    '造纸': {'industryMajor': '造纸', 'gbCodes': ['C2211', 'C2212', 'C2221']},
    '民航': {'industryMajor': '民航', 'gbCodes': ['G5631', 'G5611', 'G5612']},
}

SHEET_CONFIG = {
    '电力': ('2-1B', '2-1C', 4),
    '水泥': ('2-2B', '2-2C', 4),
    '平板玻璃': ('2-2B', '2-2C', 5),
    '钢铁': ('2-3B', '2-3C', 4),
    '铝冶炼': ('2-4B', '2-4C', 4),
    '铜冶炼': ('2-4B', '2-4C', 5),
    '石化': ('2-5B', '2-5C', 4),
    '化工': ('2-6B', '2-6C', 4),
    '造纸': ('2-7B', '2-7C', 4),
    '民航': ('2-8B', '2-8C', 4),
}

REPORT_SOURCES_NP = [
    '碳核查', '碳排放权配额实际履约情况', '环境信息披露报告', 'ESG报告',
    '可持续发展报告', '社会责任报告', '其他'
]
REPORT_SOURCES_P = [
    '经连续测量的碳排放数据', '建设或运营过程实际产生的数据', '可行性研究报告',
    '设计文件', '节能报告', '其他'
]

GRID_OPTIONS = [
    '华北电网', '东北电网', '华东电网', '华中电网', '西北电网', '南方电网', '西南电网', '全国平均'
]

FUEL_CATEGORIES = {'固体燃料', '液体燃料', '气体燃料'}
SKIP_PROCESS_AS_PAIR = {'脱硫试剂', '碳酸盐分解'}
SKIP_PROCESS_AS_AMOUNT = {'碳粉使用', '电极使用', '氧化亚氮排放', '石灰石煅烧'}


def slug(text):
    s = re.sub(r'[^\w\u4e00-\u9fff]+', '_', str(text or ''))
    s = re.sub(r'_+', '_', s).strip('_')
    return (s[:48] if s else 'item')


def find_sheet(wb, pattern, keyword):
    for name in wb.sheetnames:
        if pattern in name and keyword in name:
            return wb[name]
    return None


def parse_energy_categories(ws, value_col):
    categories = {}
    current_cat = None
    for row in range(3, ws.max_row + 1):
        cat = ws.cell(row, 1).value
        item = ws.cell(row, 2).value
        unit = ws.cell(row, 3).value
        val = ws.cell(row, value_col).value
        if cat and str(cat).startswith('*'):
            break
        if cat and str(cat).strip():
            current_cat = str(cat).strip()
        if not current_cat:
            continue
        if current_cat in ('购入电力', '购入热力'):
            continue
        item_s = str(item).strip() if item else ''
        if not item_s or item_s == '/':
            continue
        if val in (None, '/', ''):
            continue
        unit_s = str(unit).strip() if unit else ''
        bucket = categories.setdefault(current_cat, [])
        if any(x['label'] == item_s for x in bucket):
            continue
        bucket.append({
            'key': slug(f'{current_cat}_{item_s}'),
            'label': item_s,
            'unit': unit_s,
        })
    return [{'category': k, 'items': v} for k, v in categories.items()]


def build_process_blocks(categories):
    blocks = []
    for cat in categories:
        name = cat['category']
        items = cat['items']
        labels = [x['label'] for x in items]
        if name in SKIP_PROCESS_AS_PAIR:
            block_type = 'desulfur' if name == '脱硫试剂' else 'carbonate'
            blocks.append({
                'type': block_type,
                'label': name,
                'typeOptions': labels,
                'keyPrefix': slug(name),
            })
        elif '生产过程' in name:
            blocks.append({
                'type': 'process',
                'label': name.replace('：', ' — '),
                'typeOptions': labels,
                'keyPrefix': slug(name),
            })
        elif name in SKIP_PROCESS_AS_AMOUNT:
            for it in items:
                unit = it['unit'] or 't（吨）'
                blocks.append({
                    'type': 'amount',
                    'key': slug(f'{name}_{it["label"]}'),
                    'label': f'{name} · {it["label"]}（{unit}）',
                })
        elif name == '碳酸盐分解' and any('倒算' in x['label'] for x in items):
            pair_items = [x for x in items if '倒算' not in x['label']]
            amount_items = [x for x in items if '倒算' in x['label']]
            if pair_items:
                blocks.append({
                    'type': 'carbonate',
                    'label': '碳酸盐分解',
                    'typeOptions': [x['label'] for x in pair_items],
                    'keyPrefix': 'carbonate',
                })
            for it in amount_items:
                blocks.append({
                    'type': 'amount',
                    'key': slug(it['label']),
                    'label': f'{it["label"]}（{it["unit"] or "t（吨）"}）',
                })
    return blocks


def parse_products(ws, sheet_name=None):
    products = []
    current_major = ''
    for row in range(3, ws.max_row + 1):
        c1 = ws.cell(row, 1).value
        c2 = ws.cell(row, 2).value
        c3 = ws.cell(row, 3).value
        if c1 and str(c1).startswith('*'):
            break
        if c1 and str(c1).strip():
            current_major = str(c1).strip()
        sub = str(c2).strip() if c2 else ''
        unit = str(c3).strip() if c3 else '吨'
        if not current_major:
            continue
        if sub in ('/', '', '——'):
            label = current_major
        elif sub:
            label = sub
        else:
            continue
        entry = {
            'key': slug(f'{current_major}_{label}'),
            'label': f'{label}（{unit}）' if unit else label,
            'group': current_major,
            'unit': unit,
            '_major': current_major,
        }
        products.append(entry)
    if sheet_name:
        products = filter_products(products, sheet_name)
    for p in products:
        p.pop('_major', None)
    return products


def filter_products(products, sheet_name):
    if sheet_name == '水泥':
        return [p for p in products if '水泥' in p.get('group', '')]
    if sheet_name == '平板玻璃':
        return [p for p in products if '平板' in p.get('group', '')]
    if sheet_name == '铝冶炼':
        return [p for p in products if p.get('group') == '铝冶炼']
    if sheet_name == '铜冶炼':
        return [p for p in products if p.get('group') == '铜冶炼']
    return products


def build_template(wb, sheet_name, biz_type):
    meta = SHEET_META[sheet_name]
    pat, patc, value_col = SHEET_CONFIG[sheet_name]
    en_ws = find_sheet(wb, pat, '能源法')
    pr_ws = find_sheet(wb, patc, '产品法')
    if not en_ws:
        raise RuntimeError(f'未找到能源法 sheet: {pat}')

    all_categories = parse_energy_categories(en_ws, value_col)
    fuel_categories = [c for c in all_categories if c['category'] in FUEL_CATEGORIES]
    process_blocks = build_process_blocks(all_categories)

    products = parse_products(pr_ws, sheet_name) if pr_ws else []
    has_product = len(products) > 0
    is_project = biz_type == 'project'
    has_heat = sheet_name != '电力'

    energy = {
        'fuelCategories': fuel_categories,
        'gridOptions': GRID_OPTIONS,
        'gridLabel': '项目所属电网' if is_project else '企业所属电网',
        'hasPurchasedHeat': has_heat,
        'processBlocks': process_blocks,
        'allowCustomFuel': True,
    }

    return {
        'id': f'{biz_type}_{sheet_name}',
        'bizType': biz_type,
        'sheetName': sheet_name,
        'industryMajor': meta['industryMajor'],
        'gbCodes': meta['gbCodes'],
        'methods': {
            'report': {
                'sourceOptions': REPORT_SOURCES_P if is_project else REPORT_SOURCES_NP,
                'hasAttachments': True,
            },
            'energy': energy,
            'product': {
                'supported': has_product,
                'fields': products,
                'allowCustomProducts': True,
            },
        },
    }


def build_all():
    if not PBO_PATH.exists():
        raise FileNotFoundError(f'附2 Excel 不存在: {PBO_PATH}')
    wb = openpyxl.load_workbook(PBO_PATH, data_only=True)
    templates = []
    for sheet in INDUSTRY_SHEETS:
        templates.append(build_template(wb, sheet, 'non_project'))
        templates.append(build_template(wb, sheet, 'project'))
    return templates


def write_outputs(templates):
    out_dir = ROOT / 'assets' / 'data'
    out_dir.mkdir(parents=True, exist_ok=True)
    meta = {
        'version': 'pbo-annex2-2026',
        'source': '人行48号文附2八大行业碳核算信息采集表、碳排放因子表',
        'count': len(templates),
        'templates': templates,
    }
    json_path = out_dir / 'supplement-templates.json'
    json_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding='utf-8')
    js_path = ROOT / 'assets' / 'js' / 'supplement-templates-data.js'
    js_path.write_text(
        '/** 采集模板 · 由 scripts/parse-supplement-templates.py 从人行附2生成 */\n'
        f'window.SUPPLEMENT_TEMPLATES = {json.dumps(templates, ensure_ascii=False)};\n',
        encoding='utf-8'
    )
    return json_path, js_path


def main():
    templates = build_all()
    json_path, js_path = write_outputs(templates)
    print(f'Generated {len(templates)} supplement templates from {PBO_PATH.name}')
    print(f'  Wrote {json_path}')
    print(f'  Wrote {js_path}')


if __name__ == '__main__':
    main()
