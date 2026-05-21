#!/usr/bin/env python3
"""解析指引附2 Excel，生成 factors-guide.json 与 factors-guide-data.js"""
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = Path('/Users/fangdanyang/Desktop/华夏银行/人行投融资碳核算工作要求/附件5．中国人民银行信贷市场司关于开展年度金融机构投融资碳核算工作的通知（银信贷〔2026〕3号）/附件2 《操作指引》附2八大行业碳核算信息采集表、碳排放因子表.xlsx')

INDUSTRY_MAP = {
    '1': '电力', '2': '建材', '3': '钢铁', '4': '有色',
    '5': '石化', '6': '化工', '7': '造纸', '8': '民航',
}

METHOD_NAMES = {
    'energy': '物理活动法-能源法',
    'product': '物理活动法-产品法',
    'economy': '经济活动法',
}


def norm_unit(raw):
    if not raw:
        return 'tCO2e', ''
    s = str(raw).strip()
    unit_raw = s
    # 简化展示单位
    m = re.search(r'（(.+?)）', s)
    short = m.group(1) if m else s
    if '兆瓦时' in s or 'MWh' in s:
        return 'tCO2e/MWh', unit_raw
    if '万千瓦时' in s:
        return 'tCO2e/万kWh', unit_raw
    if '万立方米' in s or '104Nm3' in s or '万m' in s:
        return 'tCO2e/万m³', unit_raw
    if '人次' in s:
        return 'tCO2e/人次', unit_raw
    if '吨' in s or s.startswith('t'):
        return 'tCO2e/t', unit_raw
    if '万元' in s:
        return 'tCO2e/万元', unit_raw
    return f'tCO2e/{short}', unit_raw


def parse_value(val):
    if val is None:
        return None, 'na'
    if isinstance(val, (int, float)):
        return float(val), 'default'
    s = str(val).strip()
    if s in ('/', '—', '——', '-', ''):
        if s in ('——', '—', '-'):
            return None, 'custom'
        return None, 'na'
    try:
        return float(s), 'default'
    except ValueError:
        return None, 'custom'


def extract_sub_industry(col_header):
    if not col_header:
        return None
    s = str(col_header).replace('\n', '')
    # 单位碳排放量-水泥（tCO2e）
    m = re.search(r'单位碳排放量[-—](.+?)（', s)
    if m:
        return m.group(1).strip()
    m = re.search(r'单位碳排放量[-—](.+)$', s)
    if m:
        return m.group(1).strip()
    if '碳排放因子' in s:
        return None
    return None


def industry_from_sheet(name):
    m = re.match(r'2-(\d+)[ABC]', name)
    if m:
        return INDUSTRY_MAP.get(m.group(1))
    return None


def parse_energy_sheet(ws, sheet_code, industry):
    records = []
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return records
    header = rows[1]
    value_cols = []
    for i, h in enumerate(header):
        if i < 3:
            continue
        if h and '单位碳排放量' in str(h):
            value_cols.append((i, extract_sub_industry(h)))
        elif i == 3 and h and '单位碳排放量' in str(h):
            value_cols.append((i, None))
    if not value_cols and len(header) > 3 and header[3]:
        value_cols = [(3, None)]

    cur_category = None
    seq = 0
    footnotes = []

    for row in rows[2:]:
        if not row:
            continue
        if row[0] and str(row[0]).startswith('*'):
            footnotes.append(str(row[0]))
            continue
        if row[0]:
            cat = str(row[0]).strip()
            if not cat.startswith('*'):
                cur_category = cat
        item = row[1]
        if not item or str(item).strip() in ('', '/'):
            continue
        item_name = str(item).strip()
        unit_raw = row[2]
        unit, unit_raw_keep = norm_unit(unit_raw)

        for col_idx, sub_ind in value_cols:
            val_raw = row[col_idx] if col_idx < len(row) else None
            value, value_type = parse_value(val_raw)
            if value_type == 'na':
                continue
            seq += 1
            sub = sub_ind
            if sub is None and len(value_cols) > 1:
                sub = None
            records.append({
                'id': f'EF-{sheet_code}-{seq:03d}',
                'sourceSheet': sheet_code,
                'industryMajor': industry,
                'methodId': 'energy',
                'methodName': METHOD_NAMES['energy'],
                'energyCategory': cur_category,
                'itemName': item_name,
                'subIndustry': sub,
                'productMajor': None,
                'productSub': None,
                'gbCode': None,
                'gbIndustryName': None,
                'value': value,
                'unit': unit,
                'unitRaw': unit_raw_keep or (str(unit_raw).strip() if unit_raw else unit),
                'valueType': value_type,
                'sourceNote': footnotes[0][:120] if footnotes else '指引附2',
                'isBuiltin': True,
                'status': 'active',
            })
    return records


def parse_product_sheet(ws, sheet_code, industry):
    records = []
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return records
    cur_product = None
    seq = 0
    source_note = '指引附2'

    for row in rows[2:]:
        if not row:
            continue
        if row[0] and str(row[0]).startswith('*'):
            source_note = str(row[0]).lstrip('*').strip()
            continue
        if row[0]:
            cur_product = str(row[0]).strip()
        item = row[1]
        if not item:
            continue
        item_name = str(item).strip()
        if item_name in ('/', '—', '——'):
            continue
        unit_raw = row[2]
        unit, unit_raw_keep = norm_unit(unit_raw)
        value, value_type = parse_value(row[3] if len(row) > 3 else None)
        seq += 1
        records.append({
            'id': f'EF-{sheet_code}-{seq:03d}',
            'sourceSheet': sheet_code,
            'industryMajor': industry,
            'methodId': 'product',
            'methodName': METHOD_NAMES['product'],
            'energyCategory': None,
            'itemName': None,
            'subIndustry': None,
            'productMajor': cur_product,
            'productSub': item_name,
            'gbCode': None,
            'gbIndustryName': None,
            'value': value,
            'unit': unit,
            'unitRaw': unit_raw_keep or (str(unit_raw).strip() if unit_raw else unit),
            'valueType': value_type,
            'sourceNote': source_note,
            'isBuiltin': True,
            'status': 'active',
        })
    return records


def parse_economy_sheet(ws):
    records = []
    rows = list(ws.iter_rows(values_only=True))
    cur_industry = None
    seq = 0
    source_note = '指引附2-9 · 碳排放因子=碳排放量/营业收入'

    for row in rows[2:]:
        if not row:
            continue
        if row[0] and str(row[0]).startswith('*'):
            source_note = str(row[0]).lstrip('*').strip()
            continue
        if row[0]:
            ind = str(row[0]).strip()
            if not ind.startswith('*') and ind in INDUSTRY_MAP.values():
                cur_industry = ind
        gb_code = row[1]
        gb_name = row[2]
        if not gb_code or not gb_name:
            continue
        gb_code = str(gb_code).strip()
        gb_name = str(gb_name).strip()
        value, value_type = parse_value(row[3] if len(row) > 3 else None)
        if value_type == 'na':
            continue
        seq += 1
        records.append({
            'id': f'EF-2-9-{seq:03d}',
            'sourceSheet': '2-9',
            'industryMajor': cur_industry,
            'methodId': 'economy',
            'methodName': METHOD_NAMES['economy'],
            'energyCategory': None,
            'itemName': None,
            'subIndustry': None,
            'productMajor': None,
            'productSub': None,
            'gbCode': gb_code,
            'gbIndustryName': gb_name,
            'value': value,
            'unit': 'tCO2e/万元',
            'unitRaw': 'tCO2e/万元人民币',
            'valueType': value_type,
            'sourceNote': source_note,
            'isBuiltin': True,
            'status': 'active',
        })
    return records


def parse_workbook(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    all_records = []
    for name in wb.sheetnames:
        if 'B.能源法' in name or 'B.' in name and '能源法' in name:
            ind = industry_from_sheet(name)
            code = name.split('.')[0]
            ws = wb[name]
            all_records.extend(parse_energy_sheet(ws, code, ind))
        elif 'C.产品法' in name or ('C.' in name and '产品法' in name):
            ind = industry_from_sheet(name)
            code = name.split('.')[0]
            ws = wb[name]
            all_records.extend(parse_product_sheet(ws, code, ind))
        elif name.startswith('2-9'):
            all_records.extend(parse_economy_sheet(wb[name]))
    wb.close()
    return all_records


def write_outputs(records, out_dir):
    out_dir.mkdir(parents=True, exist_ok=True)
    json_path = out_dir / 'factors-guide.json'
    js_path = out_dir.parent / 'js' / 'factors-guide-data.js'
    meta = {
        'version': 'guide-2026-annex2',
        'source': '操作指引附2八大行业碳核算信息采集表、碳排放因子表',
        'generatedAt': __import__('datetime').datetime.now().isoformat(timespec='seconds'),
        'count': len(records),
        'factors': records,
    }
    json_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding='utf-8')
    js_body = json.dumps(records, ensure_ascii=False)
    js_path.write_text(
        '/** 指引附2内置排放因子 · 由 scripts/parse-guide-factors.py 生成，请勿手改 */\n'
        f'window.FACTORS_GUIDE = {js_body};\n',
        encoding='utf-8'
    )
    return json_path, js_path


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.exists():
        print(f'ERROR: file not found: {xlsx}', file=sys.stderr)
        sys.exit(1)
    records = parse_workbook(xlsx)
    by_method = {}
    for r in records:
        by_method[r['methodId']] = by_method.get(r['methodId'], 0) + 1
    json_path, js_path = write_outputs(records, ROOT / 'assets' / 'data')
    print(f'Parsed {len(records)} factors')
    print(f'  energy: {by_method.get("energy", 0)}')
    print(f'  product: {by_method.get("product", 0)}')
    print(f'  economy: {by_method.get("economy", 0)}')
    print(f'Wrote {json_path}')
    print(f'Wrote {js_path}')


if __name__ == '__main__':
    main()
